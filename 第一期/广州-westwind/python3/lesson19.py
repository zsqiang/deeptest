#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2018/1/20 14:09
# @Author  : caijiangchun
# @Site    : 
# @File    : lesson19.py
# @Software: PyCharm
from openpyxl import Workbook
from datetime import datetime
from openpyxl import load_workbook

def excel_write():
    print("excel简单示例")

    #创建一个Excel工作区
    wb = Workbook()

    #激活当前工作薄
    ws = wb.active

    #往单元格A1写入数据，其他单元格写入类似
    ws['A1'] = u"开源test"

    #写下一行数据，列表元素对应每一个单元格
    ws.append([1, 2, 3])

    #写入时间类型到Excel,python会自动类型转换成Excel的日期时间类型
    ws["A2"] = datetime.now()

    #保存为excel文件
    wb.save("简单的excel写的示例.xlsx")

def excel_read():

    wb = load_workbook("简单的excel写的示例.xlsx")

    #获取所有的sheet名，返回的是list类型
    sheets = wb.get_sheet_names()
    print(type(sheets))
    #print(sheets)

    #遍历sheets,并读取其单元格内容并打印输出
    for sh in sheets:
        print("读取工作簿名称： ", sh)

    #获取要读取的sheet
    ws = wb.get_sheet_by_name(sheets[0])

    #读取A1单元格的值
    A1 = ws['A1'].value
    print("A1单元格的值： ", A1)

    #读取A2, B2, C2
    for index in ('A2', 'B2', 'C2'):
        print(index, "单元格的值： ", ws[index].value)

    #读取空值的单元格，openpyxl对于空值的单元格，返回None
    F1 = ws['F1'].value
    print("F1单元格的值： ", F1)


if __name__ == "__main__":

    #excel_write()
    excel_read()