# -*- coding:utf-8 -*-
__author__ = 'Lakisha'

import datetime
from openpyxl import Workbook
if __name__ == "__main__":
    print("写excel简单示例")

    #创建一个excel工作区
    wb = Workbook()

    #激活当前工作簿
    ws = wb.active

    # 往单元格A1写入数据, 其他单元格写入类似
    ws['A1'] = "开源优测"

    ws.append([1, 2, 3])

    ws['A2'] = datetime.datetime.now()

    wb.save("简单excel写示例.xlsx")

'''
# -*- coding:utf-8 -*-
__author__ = 'Lakisha'

from openpyxl import load_workbook

if __name__ == "__main__":
    print("读取已存在的excel文件")

    wb = load_workbook("简单excel写示例.xlsx")

    # 获取所有sheet名, 返回的是list类型
    sheets = wb.get_sheet_names()
    print(type(sheets))

    for sh in sheets:
        print("读取工作簿名称： ", sh)

    # 获取要读取的sheet
    ws = wb.get_sheet_by_name(sheets[0])

    # 读取Sheet A1 ， A2, B2, C2单元格内容

    # 读取A1单元格的值
    A1 = ws["A1"].value
    print("A1单元格的值：", A1)

    # 读取A2, B2, C2
    for index in ('A2', 'B2', 'C2'):
        print(index, "单元格的值： ", ws[index].value)

    # 读取空值的单元格, openpyxl对于空值的单元格，返回None
    F1 = ws['F1'].value
    print("F1单元格的值： ", F1)
'''