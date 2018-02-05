# -*- coding:utf-8 -*-
__author__ = "Lakisha"

from openpyxl import Workbook, load_workbook

if __name__ =="__main__":
    print("python openpyxl 基本示例")

    # 创建excel文档
    wb = Workbook()
    ws = wb.active

    #给默认的工作簿修改名称
    ws.title = "我的默认创建的工作簿"

    #对第二行，A-F列写入数据
    for col in ("A", "B", "C", "D", "E", "F"):
        ws["%s2" %col] = "大数据测试"

    #对第一行A-F列写入数据
    for col in ("A", "B", "C", "D", "E", "F"):
        ws["%s1" %col] = "开源优测1"

    ws1 = wb.create_sheet("新创建的工作簿2")

    for col in ("A", "B", "C", "D", "E", "F"):
        ws1["%s1"%col] = "开源优测2"

    for col in ("A", "B", "C", "D", "E", "F"):
        ws1["%s2"%col] = "快学Python3"

    wb.save("openpyxl_demo.xlsx")

    #读取openpyxl_demo.xlsx文档中的内容
    #只读模式打开
    r_wb = load_workbook(filename="openpyxl_demo.xlsx", read_only=True)

    #获取所有工作簿名称
    sheets = r_wb.get_sheet_names()

    #遍历各个工作簿中的内容
    #即上述写入到第一、二行A-F列的数据
    for sheet in sheets:
        ws = r_wb[sheet]
        print("-" * 20)
        print(">>>读取", sheet)

        #遍历第一行，第二行A-F列的数据
        for row in (1, 2):
            for col in ("A", "B", "C", "D", "E", "F"):
                print(ws["%s%d"%(col,row)].value, end="    ")

            print(end="\n")

    #一次性读取多个单元格的数据
    for sheet in sheets:
        ws = r_wb[sheet]
        print("-" *20)
        print(">>>读取", sheet)

        cells_range=ws["A1":"F2"]
        for cells in cells_range:
            for cell in cells:
                print(cell.value, end = " ")
        print(end="\n")