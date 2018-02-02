import os,sys
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl import Workbook,load_workbook
from openpyxl.styles import colors
# 只读模式打开
wb=load_workbook(filename='openpyxl_demo.xlsx',read_only=False)
#获取所有工作簿名称
sheets=wb.get_sheet_names()
# 获取要读取的sheet
ws = wb.get_sheet_by_name(sheets[0])
# 读取A1单元格的值
A1 = ws['A1'].value
print("A1单元格的值： ", A1)
#设置单元格字体、颜色设置
a1=ws['A1']
d2=ws['D2']
a1.font = Font(color=colors.RED, italic=True)# 创建字体实例（红色、斜体，其他属性默认)
wb.save('openpyxl_demo.xlsx')
