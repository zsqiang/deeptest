from openpyxl import Workbook,load_workbook
'''

#创建excel文档
wb=Workbook()
ws=wb.active
# 给默认的工作簿修改名称
ws.title='我的默认创建的工作簿'
#对第一行,A-F列写入数据
for col in ("A", "B", "C", "D", "E", "F"):
    ws['%s1'%col]='棒棒糖'
#对第二行
for col in ("A", "B", "C", "D", "E", "F"):
    ws['%s2' % col] = '碰一个'
# 创建一个工作簿
ws1 = wb.create_sheet("新创建的工作簿2")
# 对第一行,A-F列写入数据
for col in ("A", "B", "C", "D", "E", "F"):
    ws1["%s1" % col] = "开源优测2"
# 对第二行,A-F列写入数据
for col in ("A", "B", "C", "D", "E", "F"):
    ws1["%s2" % col] = "快学Python3"
# 保存excel文档
wb.save('openpyxl_demo.xlsx')
'''

# 读取openpyxl_deml.xlsx文档中的内容
# 只读模式打开
r_wb=load_workbook(filename='openpyxl_demo.xlsx',read_only=True)
#获取所有工作簿名称
sheets=r_wb.get_sheet_names()
print(sheets)

'''
# 获取要读取的sheet
ws = wb.get_sheet_by_name(sheets[0])
# 读取A1单元格的值
A1 = ws['A1'].value
print("A1单元格的值： ", A1)
'''


# 遍历sheets，并读取其单元格内容打印输出
# 遍历各个工作簿中的内容
for sheet in sheets:
    ws=r_wb[sheet]
    print('-'*20)
    # 遍历各个工作簿的名字
    print('读取',sheet)
    '''
        for row in (1,2):
        for col in ("A", "B", "C", "D", "E", "F"):
            print(ws['%s%d'%(col,row)].value,end='  ')
        print(end='\n')
    '''
    print(ws.cell(row=1,column=1).value)
    print(ws['A2'].value)

'''
# 一次性读取多个单元格的数据
print(end='\n\n')
print("一次性读取多个单元格数据")

for sheet in sheets:
    ws=r_wb[sheet]
    print("---" * 20)
    print(">>>读取", sheet)
    cells_range=ws['A1':'F2']
    #遍历下已读取的单元格的内容
    for cells in cells_range:
        for cell in cells:
            print(cell.value,end='  ')
        print(end='\n')
'''

