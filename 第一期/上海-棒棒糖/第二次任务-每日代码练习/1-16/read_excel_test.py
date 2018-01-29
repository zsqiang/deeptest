__author__='棒棒糖'
from openpyxl import load_workbook
if __name__=='__main__':
    print('读取已经存在的excel文件')
    wb=load_workbook('excel示例.xlsx')
    #获取所有的sheet名，
    sheets=wb.get_sheet_names()
    print(type(sheets))
    #遍历sheets，并打印
    for sh in sheets:
        print('读取工作簿名称：',sh)
    #获取要读取的sheet
    ws=wb.get_sheet_by_name(sheets[0])
    #读取Sheet A1 ， A2, B2, C2单元格内容
    A1=ws['A1'].value
    print('A1=',A1)
    for index in ('A2','B2','C2'):
        print(index,'=',ws[index].value)
    #读取空值的单元格, openpyxl对于空值的单元格，返回None
    F1=ws['F1'].value
    print('F1=',F1)