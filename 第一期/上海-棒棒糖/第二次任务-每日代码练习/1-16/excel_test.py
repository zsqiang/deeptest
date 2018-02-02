__author__='棒棒糖'
from openpyxl import Workbook
if __name__=='__main__':
    print('excel简单示例')
    #创建一个excel工作区
    wb=Workbook()
    # 激活当前工作簿
    ws=wb.active
    #往单元格A1写入数据, 其他单元格写入类似
    ws['A1']='棒棒糖'
    #写下一行数据，列表元素对应每一个单元格
    ws.append([1,2,3])
    #写入时间类型到excel, python会自动将类型转换成excel的日期时间类型
    import datetime
    ws['A2']=datetime.datetime.now()
    #保存excel文件
    wb.save('excel示例.xlsx')

