# -*- coding:utf-8 -*-
__author__ = "Heather"
from openpyxl import load_workbook
if __name__ == '__main__':
    print("excel")
    wb = load_workbook("excel写示例.xlsx")

    sheets = wb.get_sheet_names()
    print(type(sheets))

    for sh in sheets:
        print("工作簿名称: ",sh)

    ws = wb.get_sheet_by_name(sheets[0])

    A1 = ws['A1'].value
    print("A1：",A1)

    for index in ('A2','B2','C2'):
        print(index,"value:",ws[index].value)

    F1 = ws['F1'].value
    print("F1单元格",F1)