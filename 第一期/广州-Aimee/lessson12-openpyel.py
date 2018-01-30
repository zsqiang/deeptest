#coding = utf-8

from openpyxl import Workbook, load_workbook

# if __name__ == "__main__":
#     print('excel简单示例')
#
#     #创建excel工作区
#     wb = Workbook()
#     #激活当前excel
#     ws = wb.active
#     #网单元格A1写入数据，其他单元格写入类似
#     ws['A1'] = "aimee"
#     ws.append([1,2,3])
#     ws.append(["aimee", "jack", "lili"])
#
#     import datetime
#     ws['D1'] = datetime.datetime.now()
#     wb.save("简单excel写示例.xlsx")

#读文件
if __name__ == "__main__":
    print("读取已存的excel文件")
    #打开文件，
    wb = load_workbook("简单excel写示例.xlsx")
    #获取所有sheet名，返回list类型
    sheets = wb.get_sheet_names()
    print(type(sheets))

    #遍历sheets
    for sh in sheets:
        print("读取工作簿名称： ",sh)

    #获取读取的sheet
    ws = wb.get_sheet_by_name(sheets[0])
    #获取单元格的内容
    A1 = ws['A1'].value
    print("A1单元格的值：",A1)

    #遍历
    for index in ( "A1",'A2',"A3","B1",'B2',"B3"):
        print(index,"单元格的值：",ws[index].value)




