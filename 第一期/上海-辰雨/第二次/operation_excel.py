#!/usr/bin/env python
# encoding: utf-8
from openpyxl import Workbook
print ('写Excel简单示例：')
#创建一个excel 的工作区：
wb = Workbook()

#激活当前的工作簿:
ws = wb.active

#往单元格AI 写入数据，其他单元格写入类似
ws['A1'] = '开源优测'

#写入一行数据，列表元素对应每一个单元格
ws.append([1,2,3])

#写入时间类型的excel，python会自动将类型转换成excel的日期类型：
import datetime
ws['B1']=datetime.datetime.now()

#保存文件为excel文件
wb.save('excel.xlsx')

from openpyxl import load_workbook
print ('读取已存在的exlce文件')
wb = load_workbook('excel.xlsx')

#获取所有的sheet名，返回的list类型
sheets = wb.get_sheet_names()
print (type(sheets))

#遍历sheets,并读取其单元格内容打印输出：
for sh in sheets:
	print (u'读取工作簿名称：',sh)

#获取要读取的sheet:
ws = wb.get_sheet_by_name(sheets[0])

#读取Sheet A1,A2,B2,C2单元格内容
A1 = ws['A1'].value
print (u'A1单元格的值：',A1)

#读取A2，B2，C2
for index in ('A2','B2','C2'):
	print (index,u'单元格的值:',ws[index].value)

#读取空值的单元格，openpyxl对干空值的单元格，返回None
F1 = ws['F1'].value
print ('F1单元格的值:',F1)