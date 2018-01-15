# -*- coding:utf-8 -*-

__author__ = 'VV'

try:
    # you can use API (xml.etree.cElementTree) of C language if you want to make speed up
    import xml.etree.cElementTree as ET
except ImportError:
    import xml.etree.ElementTree as ET

if __name__ == '__main__':
    print("Element Tree XPath demo")

    # load xml file
    tree = ET.parse("data_demo.xml")

    # get root node, print node text
    root = tree.getroot()

    # select current node, return the list of current node
    print("Select current node")
    data = root.findall(".")
    for d in data:
        print(d.tag)

    # select all country nodes
    print("Select all country nodes way 1")
    countrys = root.findall(".//country")
    for country in countrys:
        print(country.tag, " ", country.attrib["name"])

    print("Select all country nodes way 2")
    countrys = root.findall("country")
    for country in countrys:
        print(country.tag, " ", country.attrib["name"])

    print("Select the country node that the name attribute is Panama")
    countrys = root.findall(".//*[@name='Panama']")
    for country in countrys:
        print(country.tag, " ", country.attrib["name"])

    # the year node from country node that the name attribute is Panama
    print("The year node from country node that the name attribute is Panama")
    years = root.findall(".//country[@name='Panama']/year")
    for year in years:
        print(year.text)

    # select country node by index, select the first country node
    # NOTE: index start from 1
    print("Select country node by index, select the first country node")
    country = root.findall(".//country[1]")
    for c in country:
        print(c.tag, " ", c.attrib["name"])

    # select node through child node and its text
    # select country node that the child node is hdppc and its text is 59900
    # NOTE: it will return the parent node of gdppc
    print("Select node through child node and its text")
    gdppc = root.findall(".//*[gdppc='59900']")
    for gd in gdppc:
        print(gd.tag, " ", gd.attrib["name"])


import urllib.request
import csv
import codecs

if __name__ == '__main__':
    print("urllib get info from Douban demo")
    print("search key: python")

    url = "https://api.douban.com/v2/book/search?q=python"
    response = urllib.request.urlopen(url)

    # turn bytes to string
    ebook_str = response.read().decode()

    # turn string to dict
    ebook_dict = eval(ebook_str)

    #print(ebook_str)
    #print(type(ebook_str))
    count = ebook_dict["count"]
    total = ebook_dict["total"]

    with codecs.open('books.csv', 'w', 'utf-8') as csvfile:
        spamwriter = csv.writer(csvfile, delimiter=',', quotechar='|', quoting=csv.QUOTE_MINIMAL)
        spamwriter.writerow(["Title", "Author", "Summary", "Publisher", "Price"])

        # book info
        for book in ebook_dict["books"]:
            spamwriter.writerow([book["title"],
                ",".join(book["author"]),
                book["summary"],
                book["publisher"],
                book["price"]])

        print(f"Total searching {total} book info")

        # get other book info from second page
        # 这段代码采集了大量数据，容易被封IP，所以注释了
        """
        for start in range(i, int(total / count) +1):
            url = "https://api.douban.com/v2/book/search?q=python&start=%d" % start
            try:
                response = urllib.request.urlopen(url)
            except:
                print("hahaha")
                break

        # turn bytes to string
        ebook_str = response.read().decode()

        # turn string to dict
        ebook_dict = eval(ebook_str)

        # print book info
        for book in ebook_dict["books"]:
            spamwriter.writerow([book["title"],
            ",".join(book["author"]),
            book["summary"],
            book["publisher"],
            book["price"]])



import urllib.request

if __name__ == '__main__':
    print("urllib base instance")

    url = "http://www.baidu.com"

    # access Baidu
    response = urllib.request.urlopen(url)

    # print status
    print(response.status)

    # print the instruction of status
    print(response.reason)

    # print header of return request
    print(response.headers)

    # print data of return request
    print(response.read().decode("utf-8"))
"""

from openpyxl import Workbook

if __name__ == '__main__':
    print("excel simple demo")

    # create a excel file
    wb = Workbook()

    # active current excel
    ws = wb.active

    # write data to the A1 cell
    ws['A1'] = "Ukulele"

    # wrtie the next line, element of list match with every cell
    ws.append([1, 2, 3])

    # write time to excel, python will turn type to excel date type automaticly
    import datetime
    ws['A3']= datetime.datetime.now()

    # save the excel file
    wb.save("SimpleExcelDemo.xlsx")


from openpyxl import load_workbook

if __name__ == '__main__':

    print("read already exist excel file")

    wb = load_workbook("SimpleExcelDemo.xlsx")

    # get all sheets name, return list type
    sheets = wb.get_sheet_names()
    print(type(sheets))

    # traverse sheets and print
    for sh in sheets:
        print("Read sheet name: ", sh)

    # get sheet you want to read
    ws = wb.get_sheet_by_name(sheets[0])

    # read value of A1 cell
    A1 = ws['A1'].value
    print("A1 cell value is: ", A1)

    # read A2, B2, C2 and A3
    for index in ('A2', 'B2', 'C2', 'A3'):
        print(index, "cell value: ", ws[index].value)

    # read bland cell, openpyxl will return None if it's a blank cell
    F1 = ws['F1'].value
    print("F1 cell value is: ", F1)


import urllib.request
from openpyxl import Workbook

if __name__ == '__main__':
    print("Get Douban book info and write into excel demo")

    url = "https://api.douban.com/v2/book/search?q=python"
    response = urllib.request.urlopen(url)

    # turn bytes to string
    ebook_str = response.read().decode()

    # turn string to Dict
    ebook_dict = eval(ebook_str)

    # create a Workbook object
    wb = Workbook()

    # active first sheet
    ws = wb.active

    # write table title
    ws.append(["Title", "Author", "Summary", "Publisher", "Price"])

    # write book info
    for book in ebook_dict["books"]:
        ws.append([book["title"],
            ",".join(book["author"]),
            book["summary"],
            book["publisher"],
            book["price"]])

    # Save
    wb.save("ebook.xlsx")
