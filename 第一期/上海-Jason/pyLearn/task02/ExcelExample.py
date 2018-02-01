# _*_ coding : utf-8 _*_
__author__ = "Jason"

'''
什么是openpyxl
    openpyxl是一个第三方的Python Excel读写库，支持Excel2010 xlsx/xlsm/xltx/xltm文件格式。
    openpyxl提供哪些能力
        Excel的基本读写能力
        与pandas和numpy无缝链接能力
        Excel里图标管理
        Excel单元格注释管理
安装openpyxl
    pip install openpyxl
导入openpyxl模块
    要使用openpyxl对Excel进行读写，需要导入其中的Workbook类，方法如下：
        from openpyxl import Workbook
'''
from openpyxl import Workbook

if __name__ == "__main__":
    print("写Excel简单示例")

    #创建一个Excel工作区
    wb = Workbook()

    #激活当前工作簿
    ws = wb.active

    #往单元格A1写入数据，其他单元格写入类似
    ws["A1"] = "开源优测"

    #写下一行数据，列表元素对应每一个单元格
    ws.append([1,2,3])

    #写入时间类型到Excel，Python会自动将类型转换成Excel的日期时间类型
    import datetime
    ws["A2"] = datetime.datetime.now()

    #保存为Excel文件
    wb.save("简单Excel写示例.xlsx")