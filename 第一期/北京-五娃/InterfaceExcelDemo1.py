# __author__ ='wuwa'
# -*- coding: utf-8 -*-
import os

from openpyxl import load_workbook


class OPEXCEL:
    def __init__(self, path, read_only=False):
        self.wb = None
        if os.path.exists(path):
            self.path = path
            self.wb = load_workbook(self.path, read_only=read_only)
        else:
            print('%s 文件不存在' % path)
            exit(0)

    def get_cell_row(self, sheetname):
        """
        通过sheetname获取sheet的行数
        :param sheetname:
        :return:
        """
        if self.wb:
            sh = self.wb.get_sheet_by_name(sheetname)
            if sh:
                return sh.max_row
        return None

    def get_cell_col(self, sheetname):
        """
        通过sheetname获取sheet的列数
        :param sheetname: 
        :return: 
        """
        if self.wb:
            sh = self.wb.get_sheet_by_name(sheetname)
            if sh:
                return sh.max_column
        return None

    def get_sheet_name(self):
        """
        获取sheet的名字
        :return: 
        """
        if self.wb:
            return self.wb.get_sheet_names()
        return None

    def get_sheet_name_by_index(self, index):
        """
        通过索引获取sheet
        :param index:
        :return:
        """
        if self.wb:
            sheets = self.wb.get_sheet_names()
            sheet_len = len(sheets)
            if index >= 0 and index < sheet_len:
                return sheets[index]
        return None

    def create_sheet(self, title, index):
        """
        创建sheet页
        :param title:
        :param index:
        :return:
        """
        res = False
        if self.wb:
            self.wb.create_sheet(title=title, index=index)
            res = True
        return res

    def set_sheet_name(self, sheetname, name):
        """
        更改sheet名
        :param sheetname:
        :param name:
        :return:
        """
        res = False
        if self.wb:
            self.wb[sheetname] = name
            res = True
        return res

    def get_cell_value(self, sheet, row, col):
        """
        获取单元格的值
        :param sheet:
        :param row:
        :param col:
        :return:
        """

        value = None
        if self.wb:
            value = self.wb[sheet].cell(row=row, column=col).value
        return value

    def set_cell_value(self, sheet, row, col, value):
        """
        设置单元格的值
        :param sheet:
        :param row:
        :param col:
        :param value:
        :return:
        """
        res = False
        if self.wb:
            self.wb[sheet].cell(row=row, column=col).value = value
            res = True
        return res

    def save(self, path=""):
        if path != "":
            self.path = path
        if self.wb:
            self.wb.save(self.path)


if __name__ == "__main__":
    eg = OPEXCEL('openxl_demo.xlsx')
    sheet = eg.get_sheet_name()
    print(sheet)

    print("----------------------")
    for index in range(0, len(sheet)):
        print(eg.get_sheet_name_by_index(index), end=' ')
    print('\n')

    print('-----------------------')
    for sh in sheet:
        rows = eg.get_cell_row(sh)
        cols = eg.get_cell_col(sh)
        # print(rows,cols)
        for i in range(1, cols + 1):
            value = eg.get_cell_value(sh, rows, cols)
            print("[%d,%d]->%s " % (rows, cols, value), end='\t')

    print("-----------------------")
    for sh in sheet:
        cols = eg.get_cell_col(sh)
        for col in range(1, cols + 1):
            eg.set_cell_value(sh, row=1, col=7, value='hello')

    eg.save()
