# __author__ ='wuwa'
# -*- coding: utf-8 -*-
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook

if __name__ == "__main__":
    # 创建衣蛾excel工作区
    wb = Workbook()

    # 激活当前sheet页
    ws = wb.active

    # 往单元格A1中写入数据
    ws['A1'] = "Python3"

    # 写入下一行数据
    ws.append([1, 2, 3])

    # 写入时间类型到excel, python会自动将类型转换成excel的日期时间类型
    ws['A2'] = datetime.datetime.now()

    # 保存为excel文件
    wb.save("excelDemo.xlsx")

    print("读取excel的数据")
    rb = load_workbook("excelDemo.xlsx")

    # 获取exce表格的所有sheet名，返回列表类型
    sheets = rb.get_sheet_names()
    print(type(sheets))

    # 读取每个sheet页的名字
    for each_sheet in sheets:
        print(each_sheet)

    # 获取要读取的sheet
    rs = rb.get_sheet_by_name(sheets[0])
    print(rs)

    # 读取 sheet 的内容
    print(rs['A1'].value)
    # 循环输出其他单元格的内容，用index来代替A2，A3，A4

    for index in ('A2', 'B2', 'C2'):
        # print(index)
        print(rs[index].value)

    # 读取空值的单元格, openpyxl对于空值的单元格，返回None
    F1 = ws['F1'].value
    print("F1单元格的值： ", F1)
