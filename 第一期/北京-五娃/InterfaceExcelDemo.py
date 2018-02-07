# __author__ ='wuwa'
# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import colors

if __name__ == "__main__":
    # 创建excel
    wb = Workbook()
    ws = wb.active

    # 设置sheet名称
    ws.title = 'Test'
    # 对第一行的几列写入值
    for col in ('A', 'B', 'C', 'F'):
        ws["%s1" % col] = "343"

    # 对第二行的几列写入值
    for col in ('A', 'B', 'C', 'F'):
        ws["%s2" % col] = "123"

    # 创建一个新的sheet
    ws1 = wb.create_sheet("新的工作簿")

    # 对第一行的几列写入值
    for col in ('A', 'B', 'C', 'F'):
        ws1["%s1" % col] = "deeptest"

    # 对第二行的几列写入值
    for col in ('A', 'B', 'C', 'F'):
        ws1["%s2" % col] = "快学"

    # 设置单元格字体颜色
    values = ws['A2'].value
    col_value = ws['A2']
    col_value.font = Font(name=u'宋体', sz=20, color=colors.RED)

    # 保存excel
    wb.save('openxl_demo.xlsx')

    # 合并单元格
    ws.merge_cells('A2:D3')

    # 保存excel
    wb.save('openxl_demo.xlsx')

    # 读取excel
    rwb = load_workbook(filename="openxl_demo.xlsx", read_only=True)
    # 获取所有sheet
    sheets = rwb.get_sheet_names()

    # 读取单元格内容
    for row in (1, 2):
        for col in ("A", "B", "C", "D", "E", "F"):
            print(ws["%s%d" % (col, row)].value, end='      ')
        print(end='\n')

    # 一次性读取单元格内容
    for sheet in sheets:
        ws = rwb[sheet]
        cells_rage = ws['A1':'F1']
        for cells in cells_rage:
            for cell in cells:
                print(cell.value, end=" ")

        print(end='\n')
