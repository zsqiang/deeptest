#--coding：utf-8--
#本折关于利用openpyxl重新实现一个类
import os
from openpyxl import load_workbook
from openpyxl import Workbook
class Myopenpyxl:

    def __init__(self,file_path):
        if os.path.exists(file_path):
            self.wb=load_workbook(file_path)
        else:
            print("当前目录找不到该文件")

    #返回所有工作簿的名称
    def myget_sheetsname(self):
        if self.wb:      #即若存在excel并正确读取进来则为true
            sheetsname=self.wb.get_sheet_names()
            return sheetsname
        return None

    #指定worksheet若存在就返回 否则返回None
    def mysheet(self,sheetname):
        if self.wb:
            if self.wb.get_sheet_by_name(self,sheetname):
                return self.wb.get_sheet_by_name(self,sheetname)

        return None     #这里return不应该与第二个if对齐,若第一个if为false则出错


    #返回指定sheet的cell的边界范围，sheet不存在返回None
    def myboundcell(self,sheetname):
        if self.wb:
            if self.mysheet(sheetname):
                return self.mysheet(sheetname).calculate_dimension()    #关于A1：A1看源码

        return None

    #遍历返回cell的值
    def myvalue(self,sheetname,minrow,maxrow,mincol,maxcol):
        
        if self.wb:
            row_range=self.wb[sheetname].iter_rows(minrow,maxrow,mincol,maxcol)
            for row in row_range:
                for cell in row:
                    return cell.value
        return None

    #保存文件
    def mysave(self,filename):
        if filename !='' and self.wb:
            self.wb.save(filename)
        else:
            pass