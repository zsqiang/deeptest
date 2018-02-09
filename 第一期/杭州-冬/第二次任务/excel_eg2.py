#--coding:utf-8--
#本折关于用openpyxl从磁盘读取excel文件并解析
from openpyxl import load_workbook

if __name__=="__main__":
    wb=load_workbook("excel_data.xlsx")

    print("获取所有的sheet 名","-------")
    sheets_name=wb.get_sheet_names()
    print(sheets_name)

    #返回一个sheet给我
    ws=wb["Sheet"]

    #遍历多个cell并输出value值
    cell_range=ws['A1:C3']
    
    for cells in cell_range:
        for cell in cells:
            print(cell,cell.value,end='   ')