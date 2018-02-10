#--coding:utf-8--
#本折关于用openpyxl简述创建excel并保存,详情见官方文档
#openpyxl官方解释：Openpyxl is a Python library for reading and writing Excel 2010 xlsx/xlsm/xltx/xltm files

#首先明确几个概念：workbook--工作簿,可看做我们常说的一个excel文件,一个工作簿里至少有一张工作表
    #worksheet:工作表 如我们常见一个账簿里面有用各种表来分类
    #cell:单元格 一个个数据就放在里面  用row-行 column-列来定位
    #还不懂的打开你的excel操作下就懂了


#python3没有内置安装openpyxl 需要自己安装
    #window打开你的命令行 直接： python -m pip install openpyxl
    #由于是连接国外的站点下,可能导致安装失败。多试几次或更改pip至国内镜像下载安装
from openpyxl import Workbook
if __name__=="__main__":
    #首先创建一个workbook对象
    wb=Workbook()  #同时默认创建了一个worksheeet

    #默认创建的worksheet没看到,给我返回来看看
    ws=wb.active    #现在ws就是那家伙

    #任性的我就是要自己建几个worksheet
    ws2=wb.create_sheet("mysheet2")
    ws3=wb.create_sheet("mysheet3")

    #给worksheet改个名字
    ws3.title="my3"
    #把所有的worksheet报上名来
    sheets_name=wb.get_sheet_names() #list类型
    print(sheets_name)
    
    #返回worksheet
        #一个workbook可视为dict,用作为title的key可访问woksheeet
    ss=wb["mysheet2"]
    # When a worksheet is created in memory, it contains no cells. They are created when firs accessed
    #在你的workbook对象保存到磁盘成为文件前 worksheet中是没有cell的,不过cell会在你首次访问它时自动生成

    #so 3种访问cell的方法
        #worksheet可视为dict 坐标和一个cell为key-value对
    ws["A1"]     #默认值为None,在A1返回一个 cell,不存在则穿建一个
    ws["A2"]=12     #赋值的同时创建cell
    a3=ws.cell(row=3,column=1,value=13)


    #输出3个单元格的值
    print(ws["A1"].value,ws["A2"].value,ws["A3"].value,ws["A4"].value)
    
    
    #访问多个cell
    cell_range=ws['A1:B2']  #generator
    print("看看访问了那些cell")
    for cells in cell_range: #每个cell为一个tuple,其元素为每行单元格
        print(cells)

    c4=ws[4]    #tuple 创建/访问一行cell

    print("按行访问并输出")
    for row in ws.iter_rows(min_row=1,max_row=2,min_col=1,max_col=2):
        for cell in row: #和cell_range类似  iter_rows为生成器 每个row为一个tuple 元素为没行单元格 
            print(cell)
    print('按列输出并访问')
    for col in ws.iter_cols(min_col=1,max_col=2,min_row=1,max_row=2):
        for cell in col:
            print(cell)

    print(type(ws.iter_rows('A1:A3')))  #generator
   
    print(ws["A1"])
    print(ws[3])
    
    print("给cell赋值","------------")
    ws["A4"]=4
    ws["A3"].value=3
    print(ws["A4"].value,ws["A3"].value)
    
    print("给多个cell赋值","------------")
    #对第一行A-c列穿建cell并赋值
    for col in ["A","B","C"]:
        ws["%s1" %col]=11
    #对第2行A-c列穿建cell并赋值
    for col in ["A","B","C"]:
        ws["%s2" %col]=22

    #对第3行A-c列穿建cell并赋值
    for col in ["A","B","C"]:
        ws["%s3" %col]=33

    #对第4行赋值 创建第5行的cell并 赋值
    for cell in c4:
        cell.value=44
    print(ws["A4"].value)

    #遍历输出
    for row in range(1,5):
        for col in ("A","B","C"):
            print(ws["%s%d" %(col,row)].value,end=',')
        print(end='\n') #换行

    #保存到磁盘
    #wb.save('excel_data.xlsx')