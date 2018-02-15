#--coding:utf-8
#再次练习封装一个 excel 

import os
from openpyxl import Workbook
from openpyxl import load_workbook

class Myopenpyxl:
    #初始化：用来控制对象的初始化状态在实例化一个对象时自动调用   常用来设置对象属性 初始值
    def __init__(self,file_path):
        if os.path.exists(file_path):
            self.wb=load_workbook(file_path)
            self.list_cell=[]       #用来在遍历时存储cells,以便用来读取cell.value
        else:
            print("找不到给定的文件")
            exit(0)     #找不到文件则中断程序

    #返回所有工作簿的名称
    def myget_sheetsname(self):
        if self.wb: #若存在excel文件且被正确读取进来则为true
            return self.wb.get_sheet_names()
        return None
        
    #返回指定工作簿
    def myget_sheet_by_name(self,sheet_name):
        if self.wb:
            return self.wb.get_sheet_by_name(sheet_name)
        return None

    ##返回指定sheet的cell的边界范围
    def myboundcell(self,shee_tname):
        if self.wb:
            return self.wb[sheet_name].calculate_dimension()
        return None

    #遍历返回cells
    def myvalue(self,sheet_name,minrow,maxrow,mincol,maxcol):
        
        if self.wb:
            row_range=self.wb[sheet_name].iter_rows(min_row=minrow,max_row=maxrow,min_col=mincol,max_col=maxcol)
            for row in row_range:
                self.list_cell.append(row)
            return self.list_cell
        return None

    #把上面返回的cells中的各个cell.value打印出来
    def print_myvalue(self,list_cells):
        for row_cells in list_cells:
            for cell in row_cells:
                print(cell.value,end=' ')
            print(end='\n')

    #修改工作簿名 修改成功则返回true
    def set_sheet_name(self,sheet_name,name):
        status=False    #定义一个变量 用以确定是否修改成功
        if self.wb:
            self.wb[sheet_name].title=name
            status=True
        return status   #在未修改成功或wb不存在时返False

    #设置cell值
    def set_cell_value(self,sheet_name,row,col,value):
        status=False
        if self.wb:
            self.wb[sheet_name].cell(row=row,column=col,value=value)
            status=True  
        return False

    #返回cell值
    def cell_value(self,sheet_name,col_row):
        ValueError=None
        if self.wb:
            value=self.wb[sheet_name].value
        return value 

    #保存文件
    def mysave(self,filename):
        status=False
        if filename !='' and self.wb:
            self.wb.save(filename)
            status=True
        return status

if __name__=="__main__":
    ex=Myopenpyxl('excel_data.xlsx')
    sheets_name=ex.myget_sheetsname()
    #返回所有工作簿的name
    print("sheet_name",end=' ')
    print(sheets_name)
    #返回指定的工作簿
    a_sheet=ex.myget_sheet_by_name('Sheet')
    print(a_sheet)
    print(type(a_sheet))

    print("输出value")
    ex.print_myvalue(ex.myvalue('my3',1,4,1,3))