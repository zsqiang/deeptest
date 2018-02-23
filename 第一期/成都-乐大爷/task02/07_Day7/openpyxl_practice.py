# -*- coding:utf-8 -*-

__author__ = 'catleer'


import os
import urllib.request

from openpyxl import Workbook, load_workbook

def work_dir_change():

    path = __file__
    path_demo = os.path.dirname(path)
    os.chdir(path_demo)

def simple_exam():
    print("写excel简单示例")

    # 创建一个excel工作区
    wb = Workbook()

    # 激活当前工作薄
    ws = wb.active

    # 往单元格A1写入数据，其他单元格写入类似
    ws['A1'] = '开源优测'

    # 写下一行数据，列表元素对应每一个单元格
    ws.append([1,2,3])

    # 写入时间类型到excel
    import datetime
    ws['A2'] = datetime.datetime.now()

    # 保存为excel文件
    wb.save("简单excel写示例.xlsx")

def read_excel():
    print("读取已存在的excel文件")

    wb = load_workbook("简单excel写示例.xlsx")

    # 获取所有的sheet名，返回的是list类型
    sheets = wb.get_sheet_names()
    print(type(sheets))

    # 遍历sheets，并读取其单元格内容打印输出
    for sh in sheets:
        print("读取工作薄名称： ", sh)

    # 获取要读取的sheet
    ws = wb.get_sheet_by_name(sheets[0])

    # 读取A1单元格的值
    A1 = ws['A1'].value
    print("A1单元格的值： ", A1)

    # 读取A2，B2，C2
    for index in ('A2', 'B2', 'C2'):
        print(index, "单元格的值： ", ws[index].value)

    # 读取空值的单元格，返回None
    F1 = ws['F1'].value
    print("F1单元格的值： ", F1)

work_dir_change()
# simple_exam()
read_excel()