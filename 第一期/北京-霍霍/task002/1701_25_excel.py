# -*- coding:utf-8 -*-
__author__ = "huohuo"

from openpyxl import Workbook
import datetime
 
if __name__ == "__main__":
    print("写excel简单示例")

    # 创建一个excel工作区
    wb = Workbook()
    
    # 激活当前工作薄
    ws = wb.active

    # 往单元格A1写入数据，其他单元格写入类似
    ws['A1'] = "开源优测"

    # 写下一行数据，列表元素对应每一个单元格
    ws.append([1, 2, 3])

    # 写入时间类型到Excel,python会自动将类型转换成excel的日期时间类型
    ws['A2'] = datetime.datetime.now()

    # 保存为excel文件
    wb.save("简单excel写示例.xlsx")





