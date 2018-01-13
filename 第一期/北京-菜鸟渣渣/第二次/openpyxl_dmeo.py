# coding=utf-8

from openpyxl import workbook, load_workbook

if __name__ == "__main__":

    # 1.写入
    print("py.excel demo")
    # 创建一个excel工作区
    wb = workbook.Workbook()
    # 激活当前工作簿
    ws = wb.active
    # 往单元格A1写入数据, 其他单元格写入类似
    ws["A1"] = "开源优测"
    # 写下一行数据，列表元素对应每一个单元格
    ws.append([1, 2, 3])
    # 写入时间类型到excel, python会自动将类型转换成excel的日期时间类型
    import datetime

    ws["A2"] = datetime.datetime.now()

    wb.save("simple_excel_demo.xlsx")

    # 2.读取
    print("读取已经存在的excel文件")

    wb = load_workbook("simple_excel_demo.xlsx")

    sheets = wb.get_sheet_names()

    for sh in sheets:
        print("工作部名称", sh)

    ws = wb.get_sheet_by_name(sheets[0])
    A1 = ws['A1'].value
    print("A1单元格的值：---> ", A1)

    # 读取A2, B2, C2
    for index in ('A2', 'B2', 'C2'):
        print(index, "单元格的值： ", ws[index].value)

        # 读取空值的单元格, openpyxl对于空值的单元格，返回None
    F1 = ws['F1'].value
    print("F1单元格的值： ", F1)

    print(type(sheets))

    pass
