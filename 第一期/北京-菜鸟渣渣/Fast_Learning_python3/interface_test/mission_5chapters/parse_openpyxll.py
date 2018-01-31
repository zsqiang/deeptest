# coding=utf-8
import logging

logging.basicConfig(level=logging.DEBUG, format="[%(asctime)s] %(name)s:%(levelname)s: %(message)s")
'''
我们看一下excel如何应用到我们的测试场景中来，通常的应用场景有：

用于测试数据的管理维护

用于自动化测试（含UI级、接口级等等）用例管理

用于测试报告生成

下面我们介绍下使用openpyxl对excel进行读写。

什么是openpyxl

openpyxl是一个Python库读写Excel 2010 xlsx/xlsm/xltx /XLTM的库。

注意其局限性，不支持低版本的excel。

如何安装openpyxl和pillow

直接使用pip命令进行安装，如下:

pip install openpyxl
为了让openpyxl具备处理图片的能力，你还需要安装pillow，命令如下：

pip install pillow
官方文档

openpyxl官方文档如下：

https://openpyxl.readthedocs.io/en/default/
pillow官方文档如下：

http://pillow.readthedocs.io/en/4.2.x/
openpyxl基本示例

下面我们演示一个基本示例，实现一下功能：

创建一个excel文档
创建多个工作簿，往单元格中写入一些符串
一次读取其中一个或多个单元格数据
保存创建的excel文档
读取已保存的excel中的数据  


'''

from openpyxl import Workbook, load_workbook
import time

# 创建excel文档
wb = Workbook()
ws = wb.active

# 给默认的工作簿修改名称
ws.title = "我的默认创建的工作簿"

# 对第一行,A-F列写入数据
for col in ["A", "B", "C", "D", "E", "F"]:
    ws["%s1" % col] = "极微速成"

# 对第二行,A-F列写入数据
for col in ("A", "B", "C", "D", "E", "F"):
    ws["%s2" % col] = "公众号: Young"

# 第三行写入测试数据
for col in ["A", "B", "C", "D", "E", "F"]:
    ws["%s3" % col] = col + str(time.time())
# 创建一个工作簿
ws1 = wb.create_sheet("新创建的工作簿1")

# 对第一行,A-F列写入数据
for col in ("A", "B", "C", "D", "E", "F"):
    ws1["%s1" % col] = "极微速成1"

# 对第二行,A-F列写入数据
for col in ("A", "B", "C", "D", "E", "F"):
    ws1["%s2" % col] = "大数据测试"

    # 创建一个工作簿
ws1 = wb.create_sheet("新创建的工作簿2")

# 对第一行,A-F列写入数据
for col in ("A", "B", "C", "D", "E", "F"):
    ws1["%s1" % col] = "极微速成2"

# 对第二行,A-F列写入数据
for col in ("A", "B", "C", "D", "E", "F"):
    ws1["%s2" % col] = "快学Python3"

    # 保存excel文档到硬盘
# wb.save('openpyxl_demo.xlsx')

'''only readable to openpyxl_demo.xlsx'''

# 读取openpyxl_deml.xlsx文档中的内容
# 只读模式打开
r_wb = load_workbook(filename='openpyxl_demo.xlsx',
                     read_only=True)

# 获取所有工作簿名称
sheets = r_wb.get_sheet_names()

logging.info(sheets)

# 遍历各个工作簿中的内容
# 即上述写入到第一、二行A-F列的数据
'''
for sheet in sheets:
    ws = r_wb[sheet]
    print("---" * 20)
    print(">>>读取", sheet)

    # 遍历第一、二行A-F列的数据
    for row in (1, 2):
        for col in ("A", "B", "C", "D", "E", "F"):
            print(ws["%s%d" % (col, row)].value, end='      ')
        print(end='\n')
           
'''
for sheet in sheets:

    # print(sheet)
    ws = r_wb.get_sheet_by_name(sheet)
    print(">>>>>读取", sheet)
    for row in (1, 2):
        for col in ("A", "B", "C", "D", "E", "F"):
            print(ws["%s%d" % (col, row)].value, end="   ")
        print(end="\n")

'''





        # 一次性读取多个单元格的数据
print(end='\n\n')
print(">>> 一次性读取多个单元格数据")

for sheet in sheets:
    ws = r_wb[sheet]
    print("---" * 20)
    print(">>>读取", sheet)
    cells_range = ws["A1": "F2"]
    # 遍历下已读取的单元格的内容
    for cells in cells_range:

        for cell in cells:
            print(cell.value, end='  ')

    print(end='\n')

'''
print("&" * 50)
for sheet in sheets:
    ws = r_wb.get_sheet_by_name(sheet)
    print("-----------" * 20)
    print("读取", sheet)
    for cells in ws["A1":"F2"]:
        for cell in cells:
            print(cell.value, end="  ")
    print(end="\n")
