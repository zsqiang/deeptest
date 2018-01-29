__author__="大霞"
if __name__=="__main__":
    from openpyxl import load_workbook
    wb=load_workbook("简单excel写入实例.xlsx")
    #获取所有sheet名，返回的是list类型
    sheets=wb.get_sheet_names()
    print(type(sheets))

    #遍历sheets，并读取其单元格内容打印输出
    for sh in sheets:
        print("读取工作簿的名称：",sh)

    #获取要读取的sheet
    ws=wb.get_sheet_by_name(sheets[0])

    #读取sheet A1、A2、B2、C2单元格的内容
    A1=ws['A1'].value
    print("A1单元格的值：",A1)

    #读取A2、B2、C2的值
    for index in('A2','B2','C2'):
        print(index,"单元格的值：",ws[index].value)

    #读取空值的单元格，openpyxl对于空值的单元格返回none
    F1=ws['F1'].value
    print("F1单元格的值：",F1)