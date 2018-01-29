__author__="大霞"
"""
openpyxl是一个第三方的pythonexcel读写库，支持Excel2010 xlsx/xlsm/xltx/xltm文件格式
xlsx/xlsm/xltx/xltm文件格式。
openpyxl提供哪些能力？
excel的基本读写能力
与pandas和numpy无缝链接能力
excel里图表管理
excel单元格注释管理
"""
#openpyxl的excel读写功能
from openpyxl import Workbook
if __name__=="__main__":
    print("写excel简单实例")
    #创建一个工作区
    wb=Workbook()
    #激活当前工作薄
    ws=wb.active
    #往单元格中写入数据，其它单元格中写入数据类似
    ws['A1']="开源优测"
    #写下一行数据，列表元素对应每一个单元格
    ws.append([1,2,3])
    ws.append(['测试','开源'])
    #写入时间类型excel，python会自动转换成excel的日期时间类型
    import datetime
    ws['A4']=datetime.datetime.now()
    ws.append(['a','b','c'])
    #保存excel文件
    wb.save("简单excel写入实例.xlsx")
    print("读取已存在的excel文件")
